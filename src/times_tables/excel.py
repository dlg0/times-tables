"""Excel extraction utilities.

Minimal wrapper over openpyxl for reading VEDA-tagged tables.
"""

import hashlib
import warnings
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def load_workbook(path: str) -> openpyxl.Workbook:
    """Load an Excel workbook with formula evaluation.

    Args:
        path: Path to Excel workbook file

    Returns:
        openpyxl Workbook object
    """
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            category=UserWarning,
            message=".*Data Validation extension is not supported.*",
        )
        return openpyxl.load_workbook(
            path, read_only=False, data_only=True, keep_vba=False, keep_links=False
        )


def get_sheet_names(workbook: openpyxl.Workbook) -> list[str]:
    """Get list of sheet names from workbook.

    Args:
        workbook: openpyxl Workbook object

    Returns:
        List of sheet names
    """
    return workbook.sheetnames


def find_tags(sheet: Worksheet) -> list[dict]:
    """Scan sheet for VEDA tag cells (starting with ~).

    Args:
        sheet: openpyxl Worksheet object

    Returns:
        List of dicts with tag metadata:
        {row: int, col: int, value: str, cell_ref: str}
    """
    tags = []

    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            if value is None:
                continue
            cell_str = str(value).strip()
            if cell_str.startswith("~"):
                tags.append(
                    {
                        "row": r_idx,
                        "col": c_idx,
                        "value": cell_str,
                        "cell_ref": f"{get_column_letter(c_idx)}{r_idx}",
                    }
                )

    return tags


def detect_table_bounds(sheet: Worksheet, tag_row: int, tag_col: int) -> tuple[int, int] | None:
    """Detect the horizontal boundaries (start_col, end_col) of a table.

    Scans left and right from the tag position in the header row (tag_row + 1)
    to find the contiguous block of non-empty header cells.

    The tag can be anywhere in or near the header row - doesn't have to be
    directly above a header column. If there's no header at the tag column,
    we scan right to find the first header, then expand from there.

    Args:
        sheet: openpyxl Worksheet object
        tag_row: Tag row
        tag_col: Tag column

    Returns:
        (start_col, end_col) indices (1-based), or None if no headers found
    """
    header_row = tag_row + 1
    
    # Check if tag column has a header directly below it
    anchor = sheet.cell(row=header_row, column=tag_col).value
    has_header_at_tag = anchor is not None and str(anchor).strip() != ""
    
    if has_header_at_tag:
        # Standard case: tag is above a header column
        start_search = tag_col
    else:
        # Tag is to the left of headers (or orphaned)
        # Scan right to find first non-empty header
        start_search = None
        for col in range(tag_col + 1, sheet.max_column + 1):
            val = sheet.cell(row=header_row, column=col).value
            if val is not None and str(val).strip() != "":
                start_search = col
                break
        
        if start_search is None:
            # No headers found to the right of tag
            return None
    
    # Now expand left and right from start_search to find contiguous header block
    # Scan Left from start position
    curr = start_search
    while curr > 1:
        val = sheet.cell(row=header_row, column=curr - 1).value
        if val is None or str(val).strip() == "":
            break
        curr -= 1
    start_col = curr

    # Scan Right from start position
    curr = start_search
    while curr <= sheet.max_column:
        val = sheet.cell(row=header_row, column=curr + 1).value
        if val is None or str(val).strip() == "":
            break
        curr += 1
    end_col = curr

    return start_col, end_col


def read_table_range(
    sheet: Worksheet, start_row: int, start_col: int
) -> tuple[list[str], list[list[Any]]]:
    """Read table headers and data starting from tag position.

    Args:
        sheet: openpyxl Worksheet object
        start_row: Tag row (headers are at start_row + 1)
        start_col: Tag column (leftmost column)

    Returns:
        (headers, data_rows) where:
        - headers: list of header strings
        - data_rows: list of data row lists
    """
    header_row_idx = start_row + 1
    data_start_row = start_row + 2

    # Detect bounds (expand left/right)
    bounds = detect_table_bounds(sheet, start_row, start_col)
    if bounds is None:
        # No valid table (empty cell below tag)
        return [], []
    
    actual_start_col, actual_end_col = bounds

    # Read headers
    headers = []
    for col_idx in range(actual_start_col, actual_end_col + 1):
        cell_value = sheet.cell(row=header_row_idx, column=col_idx).value
        # Should not be None based on detect logic, but safe to check
        headers.append(str(cell_value).strip() if cell_value is not None else "")

    if not headers:
        return [], []

    # Handle duplicate headers by appending position suffix
    # This preserves both columns rather than losing data
    from collections import Counter

    header_counts = Counter(headers)
    duplicates = {h for h, c in header_counts.items() if c > 1}

    if duplicates:
        import logging

        logger = logging.getLogger(__name__)
        logger.warning(f"Duplicate headers found in Excel: {list(duplicates)}")

        # Make duplicates unique by appending _col<position>
        seen = {}
        unique_headers = []
        for i, h in enumerate(headers):
            if h in duplicates:
                if h not in seen:
                    seen[h] = 1
                    unique_headers.append(f"{h}_col{i}")
                else:
                    unique_headers.append(f"{h}_col{i}")
            else:
                unique_headers.append(h)

        headers = unique_headers
        logger.warning("Renamed duplicate headers with column positions")

    num_cols = len(headers)

    # Read data rows
    data_rows = []
    row_idx = data_start_row

    while row_idx <= sheet.max_row:
        # Read current row
        row_data = []
        is_empty = True

        for col_offset in range(num_cols):
            # Use actual_start_col instead of start_col (which was tag_col)
            cell_value = sheet.cell(row=row_idx, column=actual_start_col + col_offset).value
            row_data.append(cell_value)
            if cell_value is not None:
                is_empty = False

        # Stop at first completely empty row
        if is_empty:
            break

        data_rows.append(row_data)
        row_idx += 1

    return headers, data_rows


def hash_workbook(path: str) -> str:
    """Generate SHA256 hash of workbook file.

    Args:
        path: Path to Excel workbook file

    Returns:
        Hex string of SHA256 hash
    """
    file_path = Path(path)
    sha256 = hashlib.sha256()

    with file_path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)

    return sha256.hexdigest()
