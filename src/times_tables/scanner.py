"""VEDA tag scanner for workbooks.

Discovers all VEDA tables in a workbook and extracts metadata.
"""

from typing import Any

from openpyxl import Workbook

from times_tables import excel


def scan_workbook(workbook: Workbook) -> list[dict[str, Any]]:
    """Scan workbook for VEDA tags and extract table metadata.

    Args:
        workbook: openpyxl Workbook object

    Returns:
        List of table metadata dictionaries with keys:
            - tag: str - Full tag string (e.g., '~FI_T: BaseParameters')
            - tag_type: str - Normalized tag type (e.g., 'fi_t')
            - logical_name: str | None - Logical name if present
            - sheet_name: str - Sheet name containing the table
            - tag_row: int - Row number of tag (1-indexed)
            - tag_col: int - Column number of tag (1-indexed)
            - headers: list[str] - Column headers
            - row_count: int - Number of data rows

    Raises:
        FileNotFoundError: If workbook file doesn't exist
    """
    tables = []

    for sheet_name in excel.get_sheet_names(workbook):
        sheet = workbook[sheet_name]
        tags = excel.find_tags(sheet)

        # Build set of tag columns for each row to detect table boundaries
        tag_positions = {}
        for tag_info in tags:
            row = tag_info["row"]
            if row not in tag_positions:
                tag_positions[row] = set()
            tag_positions[row].add(tag_info["col"])

        for tag_info in tags:
            # Parse tag to extract tag_type and logical_name
            tag_value = tag_info["value"]
            tag_type, logical_name = _parse_tag(tag_value)

            # Skip UC_Sets - these are metadata-only, not tables
            if tag_type == "uc_sets":
                continue

            # Extract headers and data rows with boundary detection
            tag_row = tag_info["row"]
            tag_col = tag_info["col"]

            # Find other tags on same row to determine boundaries (legacy/safety check)
            other_tags_on_row = tag_positions.get(tag_row, set()) - {tag_col}

            # Try to detect table bounds (returns None if no headers found anywhere)
            bounds = excel.detect_table_bounds(sheet, tag_row, tag_col)
            
            if bounds is not None:
                # Table with headers found
                headers, data_rows = _read_table_with_boundaries(
                    sheet, start_row=tag_row, start_col=tag_col, other_tag_cols=other_tags_on_row
                )
            else:
                # Single-value tag with no tabular data (e.g., ~STARTYEAR)
                headers, data_rows = [], []

            # Build table metadata dictionary
            table_dict = {
                "tag": tag_value,
                "tag_type": tag_type,
                "logical_name": logical_name,
                "sheet_name": sheet_name,
                "tag_row": tag_row,
                "tag_col": tag_col,
                "headers": headers,
                "row_count": len(data_rows),
            }

            tables.append(table_dict)

    return tables


def _read_table_with_boundaries(
    sheet, start_row: int, start_col: int, other_tag_cols: set[int]
) -> tuple[list[str], list[list[Any]]]:
    """Read table headers and data, respecting boundaries from other tags.

    Args:
        sheet: openpyxl Worksheet object
        start_row: Tag row (headers are at start_row + 1)
        start_col: Tag column (leftmost column, may be adjusted)
        other_tag_cols: Set of column indices that have tags on the same row

    Returns:
        (headers, data_rows) where:
        - headers: list of header strings
        - data_rows: list of data row lists
    """
    header_row_idx = start_row + 1
    data_start_row = start_row + 2

    # Detect bounds (expand left/right)
    bounds = excel.detect_table_bounds(sheet, start_row, start_col)
    if bounds is None:
        # No valid table (empty cell below tag)
        return [], []
    
    actual_start_col, actual_end_col = bounds

    # Respect other tags: truncate if we crossed another tag on same row
    # This handles the case where multiple tables are adjacent without a blank column separator
    if other_tag_cols:
        # Find nearest tag to the left: max(c for c in other_tag_cols if c < tag_col)
        left_tags = [c for c in other_tag_cols if c < start_col]
        if left_tags:
            # Clip start to not include columns beyond the nearest left tag
            nearest_left = max(left_tags)
            actual_start_col = max(actual_start_col, nearest_left + 1)
        
        # Find nearest tag to the right: min(c for c in other_tag_cols if c > tag_col)
        right_tags = [c for c in other_tag_cols if c > start_col]
        if right_tags:
            # Clip end to not include columns beyond the nearest right tag
            nearest_right = min(right_tags)
            actual_end_col = min(actual_end_col, nearest_right - 1)

    # Read headers
    headers = []
    col_idx = actual_start_col
    while col_idx <= actual_end_col:
        # Stop if we've reached another table's tag column (safety)
        if col_idx in other_tag_cols and col_idx != start_col:
             # Only stop if it's NOT our own tag column
             # But wait, our tag column might be in the middle of the table now.
             pass

        cell_value = sheet.cell(row=header_row_idx, column=col_idx).value
        # detect_table_bounds ensures non-None, but check anyway
        headers.append(str(cell_value).strip() if cell_value is not None else "")
        col_idx += 1

    if not headers:
        return [], []

    num_cols = len(headers)

    # Read data rows
    data_rows = []
    row_idx = data_start_row

    while row_idx <= sheet.max_row:
        # Read current row
        row_data = []
        is_empty = True

        for col_offset in range(num_cols):
            cell_value = sheet.cell(row=row_idx, column=actual_start_col + col_offset).value
            row_data.append(cell_value)
            if cell_value is not None:
                is_empty = False

        # Stop at first completely empty row
        if is_empty:
            break

        data_rows.append(row_data)
        row_idx += 1

    return headers, data_rows


def _parse_tag(tag: str) -> tuple[str, str | None]:
    """Parse VEDA tag to extract tag type and logical name.

    Args:
        tag: Full tag string (e.g., '~FI_T: BaseParameters')

    Returns:
        Tuple of (tag_type, logical_name)
        - tag_type: Normalized lowercase tag type without ~ prefix
        - logical_name: Logical name if present, None otherwise
    """
    # Remove ~ prefix
    tag_stripped = tag.lstrip("~").strip()

    # Split on colon
    if ":" in tag_stripped:
        parts = tag_stripped.split(":", 1)
        tag_type = parts[0].strip().lower()
        logical_name = parts[1].strip() if len(parts) > 1 else None
    else:
        tag_type = tag_stripped.strip().lower()
        logical_name = None

    return tag_type, logical_name
