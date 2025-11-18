"""Excel extraction utilities.

Minimal wrapper over openpyxl for reading VEDA-tagged tables.
"""

import hashlib
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def load_workbook(path: str) -> openpyxl.Workbook:
    """Load an Excel workbook with formula evaluation.

    Args:
        path: Path to Excel workbook file

    Returns:
        openpyxl Workbook object
    """
    return openpyxl.load_workbook(path, read_only=False, data_only=True)


def get_sheet_names(workbook: openpyxl.Workbook) -> list[str]:
    """Get list of sheet names from workbook.

    Args:
        workbook: openpyxl Workbook object

    Returns:
        List of sheet names
    """
    return workbook.sheetnames


def find_tags(sheet: Worksheet) -> list[dict]:
    """Scan sheet for VEDA tag cells (starting with ~).

    Args:
        sheet: openpyxl Worksheet object

    Returns:
        List of dicts with tag metadata:
        {row: int, col: int, value: str, cell_ref: str}
    """
    tags = []

    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            if value is None:
                continue
            cell_str = str(value).strip()
            if cell_str.startswith("~"):
                tags.append(
                    {
                        "row": r_idx,
                        "col": c_idx,
                        "value": cell_str,
                        "cell_ref": f"{get_column_letter(c_idx)}{r_idx}",
                    }
                )

    return tags


def read_table_range(
    sheet: Worksheet, start_row: int, start_col: int
) -> tuple[list[str], list[list[Any]]]:
    """Read table headers and data starting from tag position.

    Args:
        sheet: openpyxl Worksheet object
        start_row: Tag row (headers are at start_row + 1)
        start_col: Tag column (leftmost column)

    Returns:
        (headers, data_rows) where:
        - headers: list of header strings
        - data_rows: list of data row lists
    """
    header_row_idx = start_row + 1
    data_start_row = start_row + 2

    # Read headers
    headers = []
    col_idx = start_col
    while True:
        cell_value = sheet.cell(row=header_row_idx, column=col_idx).value
        if cell_value is None:
            break
        headers.append(str(cell_value).strip())
        col_idx += 1

    if not headers:
        return [], []

    # Handle duplicate headers by appending position suffix
    # This preserves both columns rather than losing data
    from collections import Counter

    header_counts = Counter(headers)
    duplicates = {h for h, c in header_counts.items() if c > 1}

    if duplicates:
        import logging

        logger = logging.getLogger(__name__)
        logger.warning(f"Duplicate headers found in Excel: {list(duplicates)}")

        # Make duplicates unique by appending _col<position>
        seen = {}
        unique_headers = []
        for i, h in enumerate(headers):
            if h in duplicates:
                if h not in seen:
                    seen[h] = 1
                    unique_headers.append(f"{h}_col{i}")
                else:
                    unique_headers.append(f"{h}_col{i}")
            else:
                unique_headers.append(h)

        headers = unique_headers
        logger.warning("Renamed duplicate headers with column positions")

    num_cols = len(headers)

    # Read data rows
    data_rows = []
    row_idx = data_start_row

    while row_idx <= sheet.max_row:
        # Read current row
        row_data = []
        is_empty = True

        for col_offset in range(num_cols):
            cell_value = sheet.cell(row=row_idx, column=start_col + col_offset).value
            row_data.append(cell_value)
            if cell_value is not None:
                is_empty = False

        # Stop at first completely empty row
        if is_empty:
            break

        data_rows.append(row_data)
        row_idx += 1

    return headers, data_rows


def hash_workbook(path: str) -> str:
    """Generate SHA256 hash of workbook file.

    Args:
        path: Path to Excel workbook file

    Returns:
        Hex string of SHA256 hash
    """
    file_path = Path(path)
    sha256 = hashlib.sha256()

    with file_path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)

    return sha256.hexdigest()
