"""Unit tests for VEDA tag scanner.

Tests the scan_workbook() function that discovers all VEDA tables in a workbook
and extracts metadata (tag type, logical name, position, headers, row count).
"""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from times_tables import excel
from times_tables.scanner import scan_workbook


@pytest.fixture
def synthetic_workbook():
    """Create synthetic workbook with various tag formats."""
    wb = openpyxl.Workbook()

    # Sheet 1: Multiple tables with different formats
    sheet1 = wb.active
    sheet1.title = "MultiTable"

    # Table 1: Tag with logical name at B2
    sheet1["B2"] = "~FI_T: BaseParameters"
    sheet1["B3"] = "Region"
    sheet1["C3"] = "Process"
    sheet1["D3"] = "Year"
    sheet1["E3"] = "Value"
    sheet1["B4"] = "AUS"
    sheet1["C4"] = "COAL_PWR"
    sheet1["D4"] = 2020
    sheet1["E4"] = 100.5
    sheet1["B5"] = "AUS"
    sheet1["C5"] = "GAS_PWR"
    sheet1["D5"] = 2020
    sheet1["E5"] = 75.2

    # Table 2: Tag without logical name at B8
    sheet1["B8"] = "~FI_PROCESS"
    sheet1["B9"] = "Process"
    sheet1["C9"] = "Technology"
    sheet1["B10"] = "COAL_PWR"
    sheet1["C10"] = "STEAM"
    sheet1["B11"] = "GAS_PWR"
    sheet1["C11"] = "CCGT"
    sheet1["B12"] = "WIND_PWR"
    sheet1["C12"] = "ONSHORE"

    # Table 3: Tag at different position (F2)
    sheet1["F2"] = "~TFM_INS: Commodities"
    sheet1["F3"] = "Commodity"
    sheet1["G3"] = "Unit"
    sheet1["F4"] = "ELEC"
    sheet1["G4"] = "PJ"

    # Sheet 2: Single table
    sheet2 = wb.create_sheet("SingleTable")
    sheet2["A1"] = "~FI_COMM"
    sheet2["A2"] = "Commodity"
    sheet2["B2"] = "Description"
    sheet2["A3"] = "COAL"
    sheet2["B3"] = "Black coal"
    sheet2["A4"] = "GAS"
    sheet2["B4"] = "Natural gas"

    # Sheet 3: Empty table (tag + headers but no data)
    sheet3 = wb.create_sheet("EmptyTable")
    sheet3["A1"] = "~FI_T: EmptyData"
    sheet3["A2"] = "Column1"
    sheet3["B2"] = "Column2"

    # Sheet 4: No tags
    sheet4 = wb.create_sheet("NoTags")
    sheet4["A1"] = "Random data"
    sheet4["A2"] = "Not a tag"

    return wb


@pytest.fixture
def temp_synthetic_workbook(synthetic_workbook):
    """Save synthetic workbook to temporary file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    synthetic_workbook.save(tmp_path)
    yield tmp_path

    Path(tmp_path).unlink(missing_ok=True)


@pytest.fixture
def merged_header_workbook():
    """Create workbook with merged cells in header."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "MergedHeaders"

    sheet["A1"] = "~FI_T: MergedTest"

    # Merge B2:C2 for header
    sheet.merge_cells("B2:C2")
    sheet["B2"] = "MergedColumn"
    sheet["D2"] = "NormalColumn"

    # Data row
    sheet["B3"] = "Value1"
    sheet["C3"] = "Value2"
    sheet["D3"] = "Value3"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    wb.save(tmp_path)
    yield tmp_path

    Path(tmp_path).unlink(missing_ok=True)


def test_scan_synthetic_workbook(temp_synthetic_workbook):
    """Test scanning synthetic workbook finds all tables."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # Should find 4 tables (3 in MultiTable, 1 in SingleTable, 1 in EmptyTable)
    assert len(tables) == 5


def test_scan_workbook_returns_list(temp_synthetic_workbook):
    """Test scan_workbook returns a list."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)
    assert isinstance(tables, list)


def test_table_has_required_fields(temp_synthetic_workbook):
    """Test each table has required metadata fields."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    required_fields = {
        "tag",
        "tag_type",
        "logical_name",
        "sheet_name",
        "tag_row",
        "tag_col",
        "headers",
        "row_count",
    }

    for table in tables:
        assert set(table.keys()) >= required_fields


def test_parse_tag_with_logical_name(temp_synthetic_workbook):
    """Test parsing tag with logical name."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # Find the FI_T: BaseParameters table
    base_params = [
        t
        for t in tables
        if "FI_T" in t["tag"] and "BaseParameters" in str(t.get("logical_name", ""))
    ][0]

    assert base_params["tag"] == "~FI_T: BaseParameters"
    assert base_params["tag_type"] == "fi_t"
    assert base_params["logical_name"] == "BaseParameters"


def test_parse_tag_without_logical_name(temp_synthetic_workbook):
    """Test parsing tag without logical name."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # Find the FI_PROCESS table (no logical name)
    fi_process = [t for t in tables if t["tag"] == "~FI_PROCESS"][0]

    assert fi_process["tag_type"] == "fi_process"
    assert fi_process["logical_name"] is None


def test_parse_tag_type_normalization(temp_synthetic_workbook):
    """Test tag types are normalized to lowercase."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    for table in tables:
        assert table["tag_type"].islower()
        assert table["tag_type"] == table["tag_type"].replace("~", "").split(":")[0].strip().lower()


def test_extract_sheet_name(temp_synthetic_workbook):
    """Test sheet names are correctly captured."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    sheet_names = {t["sheet_name"] for t in tables}
    assert "MultiTable" in sheet_names
    assert "SingleTable" in sheet_names
    assert "EmptyTable" in sheet_names
    assert "NoTags" not in sheet_names  # Sheet with no tags


def test_extract_tag_position(temp_synthetic_workbook):
    """Test tag row and column positions are captured."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # Find table with tag at B2 (row 2, col 2)
    base_params = [t for t in tables if "BaseParameters" in str(t.get("logical_name", ""))][0]
    assert base_params["tag_row"] == 2
    assert base_params["tag_col"] == 2

    # Find table with tag at A1 (row 1, col 1)
    single_table = [t for t in tables if t["sheet_name"] == "SingleTable"][0]
    assert single_table["tag_row"] == 1
    assert single_table["tag_col"] == 1


def test_extract_headers(temp_synthetic_workbook):
    """Test headers are extracted correctly."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # Find FI_T: BaseParameters table
    base_params = [t for t in tables if "BaseParameters" in str(t.get("logical_name", ""))][0]

    assert isinstance(base_params["headers"], list)
    assert base_params["headers"] == ["Region", "Process", "Year", "Value"]


def test_extract_headers_types(temp_synthetic_workbook):
    """Test headers are strings."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    for table in tables:
        assert isinstance(table["headers"], list)
        for header in table["headers"]:
            assert isinstance(header, str)


def test_count_data_rows(temp_synthetic_workbook):
    """Test data row counting."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # FI_T: BaseParameters should have 2 data rows
    base_params = [t for t in tables if "BaseParameters" in str(t.get("logical_name", ""))][0]
    assert base_params["row_count"] == 2

    # FI_PROCESS should have 3 data rows
    fi_process = [t for t in tables if t["tag"] == "~FI_PROCESS"][0]
    assert fi_process["row_count"] == 3


def test_empty_table_handling(temp_synthetic_workbook):
    """Test handling of table with no data rows."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # Find EmptyData table
    empty_table = [t for t in tables if "EmptyData" in str(t.get("logical_name", ""))][0]

    assert empty_table["headers"] == ["Column1", "Column2"]
    assert empty_table["row_count"] == 0


def test_multiple_tables_same_sheet(temp_synthetic_workbook):
    """Test finding multiple tables on same sheet."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    multi_table_tables = [t for t in tables if t["sheet_name"] == "MultiTable"]
    assert len(multi_table_tables) == 3


def test_tables_at_different_positions(temp_synthetic_workbook):
    """Test tables at non-A1 positions."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # Find table at F2
    commodities = [t for t in tables if "Commodities" in str(t.get("logical_name", ""))][0]
    assert commodities["tag_row"] == 2
    assert commodities["tag_col"] == 6  # Column F


def test_sheet_with_no_tags(temp_synthetic_workbook):
    """Test sheet with no tags produces no results."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    no_tag_tables = [t for t in tables if t["sheet_name"] == "NoTags"]
    assert len(no_tag_tables) == 0


def test_scan_real_fixture():
    """Test scanning real fixture workbook."""
    fixture_path = Path(__file__).parent / "fixtures" / "sample_deck" / "VT_BaseYear.xlsx"

    if not fixture_path.exists():
        pytest.skip(f"Fixture not found: {fixture_path}")

    workbook = excel.load_workbook(str(fixture_path))
    tables = scan_workbook(workbook)

    # Should find at least one table
    assert len(tables) > 0

    # Each table should have valid metadata
    for table in tables:
        assert table["tag"].startswith("~")
        assert isinstance(table["tag_type"], str)
        assert table["tag_type"].islower()
        assert isinstance(table["sheet_name"], str)
        assert isinstance(table["tag_row"], int)
        assert isinstance(table["tag_col"], int)
        assert isinstance(table["headers"], list)
        assert isinstance(table["row_count"], int)
        assert table["row_count"] >= 0


def test_merged_header_cells(merged_header_workbook):
    """Test handling of merged cells in headers."""
    workbook = excel.load_workbook(merged_header_workbook)
    tables = scan_workbook(workbook)

    assert len(tables) == 1
    table = tables[0]

    # Should handle merged cells gracefully
    assert isinstance(table["headers"], list)
    assert len(table["headers"]) > 0


def test_tag_case_insensitive_parsing(temp_synthetic_workbook):
    """Test tag type parsing is case-insensitive."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    # All tag_types should be lowercase
    for table in tables:
        tag_type = table["tag_type"]
        assert tag_type == tag_type.lower()


def test_logical_name_whitespace_handling():
    """Test logical names handle whitespace correctly."""
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Tag with extra whitespace
    sheet["A1"] = "~FI_T:   SpacedName   "
    sheet["A2"] = "Column1"
    sheet["A3"] = "Data"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    wb.save(tmp_path)

    try:
        workbook = excel.load_workbook(tmp_path)
        tables = scan_workbook(workbook)
        assert len(tables) == 1
        # Logical name should be trimmed
        assert tables[0]["logical_name"] == "SpacedName"
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_scan_nonexistent_file():
    """Test scanning nonexistent file raises error."""
    with pytest.raises((FileNotFoundError, IOError)):
        excel.load_workbook("/nonexistent/path/file.xlsx")


def test_table_metadata_types(temp_synthetic_workbook):
    """Test all metadata fields have correct types."""
    workbook = excel.load_workbook(temp_synthetic_workbook)
    tables = scan_workbook(workbook)

    for table in tables:
        assert isinstance(table["tag"], str)
        assert isinstance(table["tag_type"], str)
        assert table["logical_name"] is None or isinstance(table["logical_name"], str)
        assert isinstance(table["sheet_name"], str)
        assert isinstance(table["tag_row"], int)
        assert isinstance(table["tag_col"], int)
        assert isinstance(table["headers"], list)
        assert isinstance(table["row_count"], int)


def test_single_column_table():
    """Test table with single column."""
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet["A1"] = "~FI_COMM"
    sheet["A2"] = "Commodity"
    sheet["A3"] = "COAL"
    sheet["A4"] = "GAS"
    sheet["A5"] = "OIL"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    wb.save(tmp_path)

    try:
        workbook = excel.load_workbook(tmp_path)
        tables = scan_workbook(workbook)
        assert len(tables) == 1
        assert tables[0]["headers"] == ["Commodity"]
        assert tables[0]["row_count"] == 3
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_wide_table():
    """Test table with many columns."""
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet["A1"] = "~FI_T: WideTable"

    # Add 10 columns
    headers = [f"Col{i}" for i in range(1, 11)]
    for i, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=i, value=header)

    # Add one data row
    for i in range(1, 11):
        sheet.cell(row=3, column=i, value=f"Value{i}")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    wb.save(tmp_path)

    try:
        workbook = excel.load_workbook(tmp_path)
        tables = scan_workbook(workbook)
        assert len(tables) == 1
        assert len(tables[0]["headers"]) == 10
        assert tables[0]["row_count"] == 1
    finally:
        Path(tmp_path).unlink(missing_ok=True)
