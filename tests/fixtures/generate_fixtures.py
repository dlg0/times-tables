#!/usr/bin/env python3
"""Generate test fixtures for VEDA workbook testing."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

FIXTURES_DIR = Path(__file__).parent
SAMPLE_DECK_DIR = FIXTURES_DIR / "sample_deck"


def create_vt_baseyear():
    """Create VT_BaseYear.xlsx with FI_PROCESS, FI_COMM, and FI_T tables."""
    wb = Workbook()

    # Sheet 1: Processes with ~FI_PROCESS
    ws_processes = wb.active
    ws_processes.title = "Processes"

    ws_processes["A1"] = "~FI_PROCESS"
    ws_processes["A1"].font = Font(bold=True)

    headers = ["TechName", "TechDesc", "Sets", "PrimaryCG", "Region"]
    for col_idx, header in enumerate(headers, start=1):
        ws_processes.cell(2, col_idx, header)

    processes_data = [
        ["ELCCOA01", "Coal power plant", "ELE", "ELC", "REG1"],
        ["ELCGAS01", "Gas power plant", "ELE", "ELC", "REG1"],
        ["ELCWIN01", "Wind turbine", "ELE,RNW", "ELC", "REG2"],
    ]
    for row_idx, row_data in enumerate(processes_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws_processes.cell(row_idx, col_idx, value)

    # Sheet 2: Commodities with ~FI_COMM
    ws_comm = wb.create_sheet("Commodities")

    ws_comm["A1"] = "~FI_COMM"
    ws_comm["A1"].font = Font(bold=True)

    headers = ["CommName", "CommDesc", "CSet", "CTSLvl"]
    for col_idx, header in enumerate(headers, start=1):
        ws_comm.cell(2, col_idx, header)

    comm_data = [
        ["ELC", "Electricity", "NRG", "PRE"],
        ["COA", "Coal", "NRG", "PRI"],
        ["GAS", "Natural Gas", "NRG", "PRI"],
    ]
    for row_idx, row_data in enumerate(comm_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws_comm.cell(row_idx, col_idx, value)

    # Sheet 3: Parameters with ~FI_T
    ws_params = wb.create_sheet("Parameters")

    ws_params["A1"] = "~FI_T: BaseParameters"
    ws_params["A1"].font = Font(bold=True)

    headers = ["Region", "TechName", "Attribute", "Commodity", "2020", "2025", "2030"]
    for col_idx, header in enumerate(headers, start=1):
        ws_params.cell(2, col_idx, header)

    params_data = [
        ["REG1", "ELCCOA01", "EFF", "ELC", 0.35, 0.36, 0.37],
        ["REG1", "ELCCOA01", "STOCK", "", 1000, "", ""],
        ["REG1", "ELCGAS01", "EFF", "ELC", 0.50, 0.52, 0.54],
        ["REG1", "ELCGAS01", "CAPACT", "", 8760, 8760, 8760],
        ["REG2", "ELCWIN01", "EFF", "ELC", 1.0, 1.0, 1.0],
        ["REG2", "ELCWIN01", "CAPACT", "", 2920, 2920, 2920],
        ["REG2", "ELCWIN01", "NCAP_COST", "", 1500, 1350, 1200],
    ]
    for row_idx, row_data in enumerate(params_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws_params.cell(row_idx, col_idx, value)

    output_path = SAMPLE_DECK_DIR / "VT_BaseYear.xlsx"
    wb.save(output_path)
    print(f"Created {output_path}")


def create_sys_settings():
    """Create SysSettings.xlsx with STARTYEAR, ENDYEAR, and regions."""
    wb = Workbook()

    # Sheet 1: Settings
    ws_settings = wb.active
    ws_settings.title = "Settings"

    ws_settings["A1"] = "~STARTYEAR"
    ws_settings["A1"].font = Font(bold=True)
    ws_settings["B1"] = 2020

    ws_settings["A3"] = "~ENDYEAR"
    ws_settings["A3"].font = Font(bold=True)
    ws_settings["B3"] = 2030

    ws_settings["A5"] = "~ACTIVEPDEF"
    ws_settings["A5"].font = Font(bold=True)
    ws_settings["B5"] = "B"

    # Sheet 2: Regions
    ws_regions = wb.create_sheet("Regions")

    ws_regions["A1"] = "~BOOK_REGIONS"
    ws_regions["A1"].font = Font(bold=True)

    ws_regions.cell(2, 1, "Region")
    ws_regions.cell(3, 1, "REG1")
    ws_regions.cell(4, 1, "REG2")

    output_path = SAMPLE_DECK_DIR / "SysSettings.xlsx"
    wb.save(output_path)
    print(f"Created {output_path}")


def create_expected_tables_json():
    """Create expected_tables.json with metadata for validation."""
    import json

    expected = {
        "description": "Expected extraction metadata for test fixtures",
        "workbooks": [
            {
                "workbook_id": "VT_BaseYear",
                "filename": "VT_BaseYear.xlsx",
                "tables": [
                    {
                        "table_id": "VT_BaseYear__Processes__FI_PROCESS",
                        "sheet_name": "Processes",
                        "tag": "~FI_PROCESS",
                        "logical_name": None,
                        "expected_columns": ["TechName", "TechDesc", "Sets", "PrimaryCG", "Region"],
                        "expected_row_count": 3,
                        "expected_primary_key": ["TechName"],
                    },
                    {
                        "table_id": "VT_BaseYear__Commodities__FI_COMM",
                        "sheet_name": "Commodities",
                        "tag": "~FI_COMM",
                        "logical_name": None,
                        "expected_columns": ["CommName", "CommDesc", "CSet", "CTSLvl"],
                        "expected_row_count": 3,
                        "expected_primary_key": ["CommName"],
                    },
                    {
                        "table_id": "VT_BaseYear__Parameters__FI_T__BaseParameters",
                        "sheet_name": "Parameters",
                        "tag": "~FI_T",
                        "logical_name": "BaseParameters",
                        "expected_columns": [
                            "Region",
                            "TechName",
                            "Attribute",
                            "Commodity",
                            "2020",
                            "2025",
                            "2030",
                        ],
                        "expected_row_count": 7,
                        "expected_primary_key": ["Region", "TechName", "Attribute", "Commodity"],
                    },
                ],
            },
            {
                "workbook_id": "SysSettings",
                "filename": "SysSettings.xlsx",
                "tables": [
                    {
                        "table_id": "SysSettings__Settings__STARTYEAR",
                        "sheet_name": "Settings",
                        "tag": "~STARTYEAR",
                        "logical_name": None,
                        "expected_columns": [],
                        "expected_row_count": 1,
                        "note": "Single-value tag, no tabular data",
                    },
                    {
                        "table_id": "SysSettings__Settings__ENDYEAR",
                        "sheet_name": "Settings",
                        "tag": "~ENDYEAR",
                        "logical_name": None,
                        "expected_columns": [],
                        "expected_row_count": 1,
                        "note": "Single-value tag, no tabular data",
                    },
                    {
                        "table_id": "SysSettings__Settings__ACTIVEPDEF",
                        "sheet_name": "Settings",
                        "tag": "~ACTIVEPDEF",
                        "logical_name": None,
                        "expected_columns": [],
                        "expected_row_count": 1,
                        "note": "Single-value tag, no tabular data",
                    },
                    {
                        "table_id": "SysSettings__Regions__BOOK_REGIONS",
                        "sheet_name": "Regions",
                        "tag": "~BOOK_REGIONS",
                        "logical_name": None,
                        "expected_columns": ["Region"],
                        "expected_row_count": 2,
                        "expected_primary_key": ["Region"],
                    },
                ],
            },
        ],
        "summary": {"total_workbooks": 2, "total_tables": 7, "total_data_rows": 15},
    }

    output_path = FIXTURES_DIR / "expected_tables.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(expected, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Created {output_path}")


def create_readme():
    """Create README.md documenting the fixtures."""
    content = """# Test Fixtures

This directory contains minimal but representative VEDA workbook fixtures for testing.

## Directory Structure

```
tests/fixtures/
├── sample_deck/          # Main test deck directory
│   ├── VT_BaseYear.xlsx  # Base year data with processes, commodities, parameters
│   └── SysSettings.xlsx  # System settings (start/end year, regions)
├── expected_tables.json  # Expected extraction metadata
├── generate_fixtures.py  # Script to regenerate fixtures
└── README.md            # This file
```

## Fixture Contents

### VT_BaseYear.xlsx

**Sheet: Processes**
- Tag: `~FI_PROCESS`
- Columns: TechName, TechDesc, Sets, PrimaryCG, Region
- Rows: 3 processes (coal, gas, wind)
- Tests: Basic process extraction, primary key sorting

**Sheet: Commodities**
- Tag: `~FI_COMM`
- Columns: CommName, CommDesc, CSet, CTSLvl
- Rows: 3 commodities (electricity, coal, gas)
- Tests: Commodity extraction, simple primary keys

**Sheet: Parameters**
- Tag: `~FI_T: BaseParameters`
- Columns: Region, TechName, Attribute, Commodity, 2020, 2025, 2030
- Rows: 7 parameter rows across 2 regions
- Tests: Composite primary keys, year columns, logical names

### SysSettings.xlsx

**Sheet: Settings**
- Tags: `~STARTYEAR`, `~ENDYEAR`, `~ACTIVEPDEF`
- Single-value tags (not tabular)
- Tests: Non-tabular tag handling

**Sheet: Regions**
- Tag: `~BOOK_REGIONS`
- Columns: Region
- Rows: 2 regions (REG1, REG2)
- Tests: Simple list extraction

## Test Coverage

These fixtures enable testing:
- ✅ Multi-table extraction from single workbook
- ✅ Multiple workbooks in same deck
- ✅ Logical names in tags (`~FI_T: BaseParameters`)
- ✅ Composite primary keys (Region + TechName + Attribute + Commodity)
- ✅ Year columns (2020, 2025, 2030)
- ✅ Single-value vs. tabular tags
- ✅ Cross-region data sorting
- ✅ Deterministic CSV output

## Data Characteristics

- **Regions**: REG1, REG2
- **Years**: 2020, 2025, 2030
- **Total rows**: ~15 data rows across all tables
- **File size**: <20KB total (minimal for fast tests)

## Regenerating Fixtures

To regenerate all fixtures:

```bash
python tests/fixtures/generate_fixtures.py
```

This will overwrite existing fixtures with clean, deterministic versions.

## Validation Reference

See `expected_tables.json` for:
- Expected table IDs
- Expected column names
- Expected row counts
- Expected primary keys

Use this for assertions in unit tests.
"""

    output_path = FIXTURES_DIR / "README.md"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"Created {output_path}")


def main():
    """Generate all test fixtures."""
    SAMPLE_DECK_DIR.mkdir(parents=True, exist_ok=True)

    print("Generating test fixtures...")
    create_vt_baseyear()
    create_sys_settings()
    create_expected_tables_json()
    create_readme()
    print("\n✅ All fixtures generated successfully!")


if __name__ == "__main__":
    main()
