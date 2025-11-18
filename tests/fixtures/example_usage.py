#!/usr/bin/env python3
"""Example usage of test fixtures in unit tests."""

from openpyxl import load_workbook

from tests.fixtures import (
    VT_BASEYEAR_PATH,
    get_all_sample_workbooks,
    get_expected_tables,
)


def example_load_expected_metadata():
    """Example: Load and use expected table metadata."""
    expected = get_expected_tables()

    print("Example 1: Load expected metadata")
    print(f"  Total workbooks: {expected['summary']['total_workbooks']}")
    print(f"  Total tables: {expected['summary']['total_tables']}")

    # Get expected columns for FI_T table
    vt_baseyear = expected["workbooks"][0]
    fi_t_table = vt_baseyear["tables"][2]
    print("\n  FI_T table:")
    print(f"    Table ID: {fi_t_table['table_id']}")
    print(f"    Expected columns: {fi_t_table['expected_columns']}")
    print(f"    Expected PK: {fi_t_table['expected_primary_key']}")
    print(f"    Expected rows: {fi_t_table['expected_row_count']}")


def example_load_workbook():
    """Example: Load and inspect a workbook."""
    print("\n\nExample 2: Load workbook")

    wb = load_workbook(VT_BASEYEAR_PATH)
    print(f"  Loaded: {VT_BASEYEAR_PATH.name}")
    print(f"  Sheets: {wb.sheetnames}")

    # Read FI_PROCESS tag
    ws = wb["Processes"]
    tag_cell = ws["A1"].value
    headers = [cell.value for cell in ws[2]]

    print("\n  Sheet 'Processes':")
    print(f"    Tag: {tag_cell}")
    print(f"    Headers: {headers}")

    # Read first data row
    first_row = [cell.value for cell in ws[3]]
    print(f"    First row: {first_row}")

    wb.close()


def example_iterate_workbooks():
    """Example: Iterate over all workbooks in test deck."""
    print("\n\nExample 3: Iterate all workbooks")

    for wb_path in get_all_sample_workbooks():
        wb = load_workbook(wb_path)
        print(f"  📘 {wb_path.name}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Find tags in first column
            tags = []
            for row_idx in range(1, min(11, ws.max_row + 1)):
                cell_value = ws.cell(row_idx, 1).value
                if cell_value and isinstance(cell_value, str) and cell_value.startswith("~"):
                    tags.append(cell_value)

            if tags:
                print(f"    📄 {sheet_name}: {', '.join(tags)}")

        wb.close()


def example_use_in_pytest():
    """Example: How to use fixtures in pytest tests."""
    print("\n\nExample 4: Usage in pytest")
    print("""
def test_extract_fi_process():
    '''Test extraction of FI_PROCESS table.'''
    from tests.fixtures import VT_BASEYEAR_PATH, get_expected_tables

    # Load expected metadata
    expected = get_expected_tables()
    fi_process_expected = expected['workbooks'][0]['tables'][0]

    # Extract table (using your extraction code)
    # table = extract_table(VT_BASEYEAR_PATH, "Processes")

    # Assertions
    # assert table.tag == "~FI_PROCESS"
    # assert len(table.columns) == len(fi_process_expected['expected_columns'])
    # assert len(table.rows) == fi_process_expected['expected_row_count']
    # assert table.primary_key == fi_process_expected['expected_primary_key']

def test_deterministic_csv_output():
    '''Test CSV output is deterministic.'''
    from tests.fixtures import VT_BASEYEAR_PATH
    import hashlib

    # Extract and format twice
    # csv1 = extract_and_format(VT_BASEYEAR_PATH)
    # csv2 = extract_and_format(VT_BASEYEAR_PATH)

    # Should be byte-for-byte identical
    # assert hashlib.sha256(csv1).hexdigest() == hashlib.sha256(csv2).hexdigest()
""")


if __name__ == "__main__":
    example_load_expected_metadata()
    example_load_workbook()
    example_iterate_workbooks()
    example_use_in_pytest()
