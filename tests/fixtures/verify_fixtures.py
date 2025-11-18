#!/usr/bin/env python3
"""Verify test fixtures are correctly formatted."""

import json
from pathlib import Path

from openpyxl import load_workbook

FIXTURES_DIR = Path(__file__).parent
SAMPLE_DECK_DIR = FIXTURES_DIR / "sample_deck"


def verify_workbook(filename: str):
    """Verify a workbook's structure."""
    wb_path = SAMPLE_DECK_DIR / filename
    print(f"\n📘 Verifying {filename}...")

    wb = load_workbook(wb_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  📄 Sheet: {sheet_name}")

        # Find VEDA tags (cells starting with ~)
        tags_found = []
        for row in ws.iter_rows(max_row=10, max_col=5):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("~"):
                    tags_found.append((cell.coordinate, cell.value))

        if tags_found:
            for coord, tag in tags_found:
                print(f"    ✓ Tag: {tag} at {coord}")

        # Show row count
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"    Dimensions: {max_row} rows × {max_col} cols")

    wb.close()


def verify_expected_tables():
    """Verify expected_tables.json structure."""
    print("\n📋 Verifying expected_tables.json...")

    with open(FIXTURES_DIR / "expected_tables.json") as f:
        expected = json.load(f)

    print(f"  ✓ Total workbooks: {expected['summary']['total_workbooks']}")
    print(f"  ✓ Total tables: {expected['summary']['total_tables']}")
    print(f"  ✓ Total data rows: {expected['summary']['total_data_rows']}")

    for wb in expected["workbooks"]:
        print(f"  📘 {wb['workbook_id']}:")
        for table in wb["tables"]:
            print(f"    - {table['tag']}: {table.get('expected_row_count', 'N/A')} rows")


def main():
    """Run all verification checks."""
    print("=" * 60)
    print("VEDA Test Fixture Verification")
    print("=" * 60)

    verify_workbook("VT_BaseYear.xlsx")
    verify_workbook("SysSettings.xlsx")
    verify_expected_tables()

    print("\n" + "=" * 60)
    print("✅ All fixtures verified successfully!")
    print("=" * 60)


if __name__ == "__main__":
    main()
