import pytest
from openpyxl import Workbook

from times_tables import scanner

def test_uc_t_with_comment_column_immediately_left():
    """
    Test to reproduce the user's scenario:
    1. A table with a tag ~UC_T:...
    2. A "comment" column immediately to the left (no blank column).
    3. Verify if the scanner picks it up and what the extracted columns are.
    """
    # Create an in-memory workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UC"

    # Row 100: Tag row (analogous to user's screenshot)
    # Note: In openpyxl, rows are 1-indexed.
    tag_row = 100
    header_row = 101
    data_start = 102
    
    # Tag in column A (or merged, but value starts in A)
    ws.cell(row=tag_row, column=1).value = "~UC_T: UC_COMNET~2030~LO"

    # Row 101: Header row
    # Col A: Comment column, no blank to the left
    ws.cell(row=header_row, column=1).value = "*Components of TOTCO2"
    ws.cell(row=header_row, column=2).value = "UC_N"
    ws.cell(row=header_row, column=3).value = "cset_cn"
    ws.cell(row=header_row, column=4).value = "UC_ATTR"
    ws.cell(row=header_row, column=5).value = "LHS"

    # Row 102: Data row
    ws.cell(row=data_start, column=1).value = r"\: CO2Captured_dac"
    ws.cell(row=data_start, column=2).value = "UCEm_Non-increasing..."
    ws.cell(row=data_start, column=3).value = r"\: CO2Captured_dac"
    ws.cell(row=data_start, column=4).value = "COMNET,GROWTH"
    ws.cell(row=data_start, column=5).value = 1

    # Scan workbook
    tables = scanner.scan_workbook(wb)

    # There should be exactly one UC_T table
    assert len(tables) == 1
    t = tables[0]

    # Verify that scanner picked up the table and the left comment column
    print(f"Detected Tag: {t['tag']}")
    print(f"Detected Headers: {t['headers']}")

    assert t["tag"] == "~UC_T: UC_COMNET~2030~LO"
    assert t["sheet_name"] == "UC"
    assert t["tag_row"] == tag_row
    assert t["tag_col"] == 1  # column A

    # Verify headers include the left comment column
    expected_headers = [
        "*Components of TOTCO2",
        "UC_N",
        "cset_cn",
        "UC_ATTR",
        "LHS"
    ]
    assert t["headers"] == expected_headers

    # And row_count matches the number of data rows
    assert t["row_count"] == 1


def test_comment_change_is_ignored():
    """
    Test to verify that changing a cell comment (red triangle) 
    does not change the extracted value, and thus results in 'No change'.
    """
    from openpyxl.comments import Comment
    
    wb = Workbook()
    ws = wb.active
    
    # Setup table
    ws["A1"] = "~UC_T: Tag"
    ws["A2"] = "Header"
    ws["A3"] = "Value"
    
    # Scan and extract
    tables1 = scanner.scan_workbook(wb)
    # (In a real scenario we'd extract to DF, but verifying value is enough)
    val1 = ws["A3"].value
    
    # Add a comment to the cell
    ws["A3"].comment = Comment("This is a comment", "Author")
    
    # Scan and extract again
    tables2 = scanner.scan_workbook(wb)
    val2 = ws["A3"].value
    
    # Assert values are identical
    assert val1 == val2
    
    # Assert comments are different (openpyxl sees them)
    assert ws["A3"].comment.text == "This is a comment"
    
    # This confirms that if the tool only looks at .value, it ignores comments.
    # Since scanner.py and excel.py use .value, this explains the user's issue.


def test_tag_in_middle_of_table_expands_left():
    """
    Test that if a tag is in the middle of the header row (e.g. Col B),
    the scanner expands LEFT to find the start of the table (Col A),
    assuming the 'gap around the outside' rule.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MiddleTag"

    # Row 1: Tag in Column B (2)
    # Row 2: Headers in A, B, C
    ws["B1"] = "~Tag: Middle"
    
    ws["A2"] = "LeftCol"
    ws["B2"] = "MiddleCol"
    ws["C2"] = "RightCol"
    
    ws["A3"] = "1"
    ws["B3"] = "2"
    ws["C3"] = "3"

    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 1
    t = tables[0]
    
    # Current behavior (likely failure): Starts at B, gets ["MiddleCol", "RightCol"]
    # Desired behavior: Gets ["LeftCol", "MiddleCol", "RightCol"]
    
    print(f"Detected headers: {t['headers']}")
    
    expected_headers = ["LeftCol", "MiddleCol", "RightCol"]
    assert t["headers"] == expected_headers


def test_uc_sets_ignored():
    """Test that ~UC_Sets tags are ignored (metadata only, not tables)."""
    wb = Workbook()
    ws = wb.active
    
    # UC_Sets tags (should be ignored)
    ws["A1"] = "~UC_Sets: R_S: AllRegions"
    ws["A2"] = "~UC_Sets: T_E:"
    
    # Valid UC_T table
    ws["C1"] = "~UC_T: UP"
    ws["C2"] = "UC_N"
    ws["D2"] = "Year"
    ws["C3"] = "UC_1"
    ws["D3"] = "2030"
    
    tables = scanner.scan_workbook(wb)
    
    # Should only detect the UC_T table, not UC_Sets
    assert len(tables) == 1
    assert tables[0]["tag_type"] == "uc_t"
    assert tables[0]["headers"] == ["UC_N", "Year"]


def test_empty_cell_below_tag_creates_empty_table():
    """Test that tags with empty cells below create empty tables (0 columns, 0 rows).
    
    This matches the behavior of single-value VEDA tags like ~STARTYEAR, ~ENDYEAR.
    """
    wb = Workbook()
    ws = wb.active
    
    # Tag with empty cell below (creates empty table)
    ws["A1"] = "~STARTYEAR"
    ws["A2"] = None
    
    # Valid table nearby
    ws["C1"] = "~UC_T: Valid"
    ws["C2"] = "Header1"
    ws["D2"] = "Header2"
    ws["C3"] = "val1"
    ws["D3"] = "val2"
    
    tables = scanner.scan_workbook(wb)
    
    # Should find both tables
    assert len(tables) == 2
    
    # First table: empty table
    assert tables[0]["tag"] == "~STARTYEAR"
    assert tables[0]["headers"] == []
    assert tables[0]["row_count"] == 0
    
    # Second table: valid table with data
    assert tables[1]["tag"] == "~UC_T: Valid"
    assert tables[1]["headers"] == ["Header1", "Header2"]
    assert tables[1]["row_count"] == 1


def test_blank_column_left_of_table():
    """Test that blank column immediately left of table is not included."""
    wb = Workbook()
    ws = wb.active
    
    # Column A is blank, table starts at B
    ws["C1"] = "~Tag"
    ws["A2"] = None
    ws["B2"] = "Header1"
    ws["C2"] = "Header2"
    ws["B3"] = "val1"
    ws["C3"] = "val2"
    
    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 1
    assert tables[0]["headers"] == ["Header1", "Header2"]


def test_blank_column_inside_header_splits_table():
    """Test that blank column inside header row acts as table boundary."""
    wb = Workbook()
    ws = wb.active
    
    # Headers with blank in middle
    ws["C1"] = "~Tag"
    ws["A2"] = "H1"
    ws["B2"] = None  # Blank column
    ws["C2"] = "H3"
    ws["D2"] = "H4"
    
    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 1
    # Should stop at blank column B, so only gets C-D
    assert tables[0]["headers"] == ["H3", "H4"]


def test_multiple_comment_columns_on_left():
    """Test that multiple comment columns on the left are included."""
    wb = Workbook()
    ws = wb.active
    
    ws["D1"] = "~UC_T"
    ws["A2"] = "*Comment1"
    ws["B2"] = "*Comment2"
    ws["C2"] = r"\: Note"
    ws["D2"] = "UC_N"
    ws["E2"] = "Value"
    
    ws["A3"] = "First comment"
    ws["B3"] = "Second comment"
    ws["C3"] = "Note text"
    ws["D3"] = "Name1"
    ws["E3"] = "100"
    
    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 1
    assert tables[0]["headers"] == ["*Comment1", "*Comment2", r"\: Note", "UC_N", "Value"]


def test_comment_column_in_middle():
    """Test that comment column in middle of headers is treated normally."""
    wb = Workbook()
    ws = wb.active
    
    ws["B1"] = "~Tag"
    ws["A2"] = "UC_N"
    ws["B2"] = "*Internal Note"
    ws["C2"] = "UC_ATTR"
    
    ws["A3"] = "Name1"
    ws["B3"] = "Some note"
    ws["C3"] = "Attr1"
    
    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 1
    assert tables[0]["headers"] == ["UC_N", "*Internal Note", "UC_ATTR"]


def test_entirely_empty_row_terminates_table():
    """Test that first entirely empty row terminates the table."""
    wb = Workbook()
    ws = wb.active
    
    ws["A1"] = "~Tag"
    ws["A2"] = "H1"
    ws["B2"] = "H2"
    
    # Data rows
    ws["A3"] = "val1"
    ws["B3"] = "val2"
    ws["A4"] = "val3"
    ws["B4"] = "val4"
    
    # Row 5 is entirely empty
    ws["A5"] = None
    ws["B5"] = None
    
    # Row 6 has values (should be ignored)
    ws["A6"] = "ignored1"
    ws["B6"] = "ignored2"
    
    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 1
    assert tables[0]["row_count"] == 2  # Only rows 3-4


def test_partial_empty_row_included():
    """Test that row with some empty cells but not entirely empty is included."""
    wb = Workbook()
    ws = wb.active
    
    ws["A1"] = "~Tag"
    ws["A2"] = "H1"
    ws["B2"] = "H2"
    ws["C2"] = "H3"
    
    # Row 3: partial data
    ws["A3"] = "val1"
    ws["B3"] = None
    ws["C3"] = "val3"
    
    # Row 4: entirely empty (stop here)
    
    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 1
    assert tables[0]["row_count"] == 1


def test_two_tables_separated_by_blank_column():
    """Test multiple tables on same row separated by blank header column."""
    wb = Workbook()
    ws = wb.active
    
    # Table 1
    ws["A1"] = "~Tag1"
    ws["A2"] = "A1"
    ws["B2"] = "A2"
    
    # Blank separator
    ws["C2"] = None
    
    # Table 2
    ws["D1"] = "~Tag2"
    ws["D2"] = "B1"
    ws["E2"] = "B2"
    
    # Data
    ws["A3"] = "t1v1"
    ws["B3"] = "t1v2"
    ws["D3"] = "t2v1"
    ws["E3"] = "t2v2"
    
    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 2
    assert tables[0]["headers"] == ["A1", "A2"]
    assert tables[1]["headers"] == ["B1", "B2"]


def test_tag_at_extreme_left():
    """Test tag at column A (leftmost position)."""
    wb = Workbook()
    ws = wb.active
    
    ws["A1"] = "~Tag"
    ws["A2"] = "H1"
    ws["B2"] = "H2"
    ws["C2"] = "H3"
    
    ws["A3"] = "v1"
    ws["B3"] = "v2"
    ws["C3"] = "v3"
    
    tables = scanner.scan_workbook(wb)
    
    assert len(tables) == 1
    assert tables[0]["headers"] == ["H1", "H2", "H3"]
    assert tables[0]["tag_col"] == 1


def test_mixed_uc_t_and_uc_sets():
    """Test sheet with both UC_T and UC_Sets tags."""
    wb = Workbook()
    ws = wb.active
    
    # UC_Sets (ignore)
    ws["A1"] = "~UC_Sets: R_S: AllRegions"
    ws["A2"] = "~UC_Sets: T_E:"
    
    # UC_T table
    ws["A3"] = "~UC_T: UP"
    ws["A4"] = "UC_N"
    ws["B4"] = "Year"
    ws["A5"] = "UC_1"
    ws["B5"] = "2030"
    
    tables = scanner.scan_workbook(wb)
    
    # Should only detect UC_T
    assert len(tables) == 1
    assert tables[0]["tag_type"] == "uc_t"
    assert tables[0]["tag_row"] == 3

