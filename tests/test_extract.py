"""Unit tests for table extraction to DataFrames.

Tests the extract_table() function that extracts VEDA tables from Excel
workbooks to pandas DataFrames with normalized column names and values.
"""

import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from times_tables import excel
from times_tables.extract import extract_table
from times_tables.scanner import scan_workbook
from times_tables.veda import VedaSchema


@pytest.fixture
def schema():
    """Load VEDA schema from vendored veda-tags.json."""
    return VedaSchema()


@pytest.fixture
def basic_workbook():
    """Create basic workbook with a simple table."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "BasicTable"

    # ~FI_T tag with standard columns
    sheet["B2"] = "~FI_T: TestTable"
    sheet["B3"] = "Region"
    sheet["C3"] = "Process"
    sheet["D3"] = "Year"
    sheet["E3"] = "Value"

    # Data rows
    sheet["B4"] = "AUS"
    sheet["C4"] = "COAL_PWR"
    sheet["D4"] = 2020
    sheet["E4"] = 100.5

    sheet["B5"] = "AUS"
    sheet["C5"] = "GAS_PWR"
    sheet["D5"] = 2025
    sheet["E5"] = 75.2

    sheet["B6"] = "NZ"
    sheet["C6"] = "WIND_PWR"
    sheet["D6"] = 2030
    sheet["E6"] = 50.0

    return wb


@pytest.fixture
def alias_workbook():
    """Create workbook with alias column names."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "AliasTable"

    # ~FI_T tag with alias columns (if schema has aliases)
    sheet["A1"] = "~FI_T: AliasTest"
    sheet["A2"] = "Region"  # Standard name
    sheet["B2"] = "Pset"  # Alias for Process (if in schema)
    sheet["C2"] = "Year"  # Standard name
    sheet["D2"] = "Value"  # Standard name

    # Data
    sheet["A3"] = "AUS"
    sheet["B3"] = "COAL_PWR"
    sheet["C3"] = 2020
    sheet["D3"] = 100.0

    return wb


@pytest.fixture
def empty_cells_workbook():
    """Create workbook with empty cells and mixed types."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "EmptyCells"

    sheet["A1"] = "~FI_T: EmptyTest"
    sheet["A2"] = "Region"
    sheet["B2"] = "Process"
    sheet["C2"] = "Year"
    sheet["D2"] = "Value"

    # Row with some empty cells
    sheet["A3"] = "AUS"
    sheet["B3"] = "COAL_PWR"
    sheet["C3"] = None  # Empty
    sheet["D3"] = 100.0

    # Row with whitespace
    sheet["A4"] = "  NZ  "
    sheet["B4"] = None
    sheet["C4"] = 2025
    sheet["D4"] = "  50.0  "

    # Row with all values
    sheet["A5"] = "USA"
    sheet["B5"] = "GAS_PWR"
    sheet["C5"] = 2030
    sheet["D5"] = 75.5

    return wb


@pytest.fixture
def multiple_tables_workbook():
    """Create workbook with multiple tables on different sheets."""
    wb = openpyxl.Workbook()

    # Sheet 1: First table
    sheet1 = wb.active
    sheet1.title = "Table1"
    sheet1["A1"] = "~FI_T: FirstTable"
    sheet1["A2"] = "Region"
    sheet1["B2"] = "Process"
    sheet1["A3"] = "AUS"
    sheet1["B3"] = "COAL_PWR"
    sheet1["A4"] = "NZ"
    sheet1["B4"] = "WIND_PWR"

    # Sheet 2: Second table
    sheet2 = wb.create_sheet("Table2")
    sheet2["C5"] = "~FI_PROCESS"
    sheet2["C6"] = "Process"
    sheet2["D6"] = "Technology"
    sheet2["C7"] = "COAL_PWR"
    sheet2["D7"] = "STEAM"
    sheet2["C8"] = "GAS_PWR"
    sheet2["D8"] = "CCGT"

    # Sheet 3: Third table
    sheet3 = wb.create_sheet("Table3")
    sheet3["B10"] = "~TFM_INS: Commodities"
    sheet3["B11"] = "Commodity"
    sheet3["C11"] = "Unit"
    sheet3["B12"] = "ELEC"
    sheet3["C12"] = "PJ"

    return wb


def save_temp_workbook(wb: openpyxl.Workbook) -> str:
    """Save workbook to temporary file and return path."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    return tmp_path


def test_extract_table_basic(schema):
    """Test basic table extraction with standard columns."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test"

    sheet["A1"] = "~FI_T: Basic"
    sheet["A2"] = "Region"
    sheet["B2"] = "Process"
    sheet["C2"] = "Year"
    sheet["A3"] = "AUS"
    sheet["B3"] = "COAL_PWR"
    sheet["C3"] = 2020
    sheet["A4"] = "NZ"
    sheet["B4"] = "WIND_PWR"
    sheet["C4"] = 2025

    wb_path = save_temp_workbook(wb)

    try:
        # Load workbook once
        workbook = excel.load_workbook(wb_path)

        # Scan to get table metadata
        tables = scan_workbook(workbook)
        assert len(tables) == 1

        # Extract table
        df = extract_table(workbook, tables[0], schema)

        # Verify shape
        assert df.shape == (2, 3)

        # Verify columns (should be normalized to lowercase canonical names)
        assert list(df.columns) == ["region", "process", "year"]

        # Verify values (should be strings)
        assert df.iloc[0]["region"] == "AUS"
        assert df.iloc[0]["process"] == "COAL_PWR"
        assert df.iloc[0]["year"] == "2020"

        assert df.iloc[1]["region"] == "NZ"
        assert df.iloc[1]["process"] == "WIND_PWR"
        assert df.iloc[1]["year"] == "2025"

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_normalizes_column_aliases(schema):
    """Test that column aliases are resolved to canonical names."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Aliases"

    # Use mixed case and potential aliases
    sheet["A1"] = "~FI_T: AliasTest"
    sheet["A2"] = "REGION"  # Different case
    sheet["B2"] = "Process"
    sheet["A3"] = "AUS"
    sheet["B3"] = "COAL_PWR"

    wb_path = save_temp_workbook(wb)

    try:
        workbook = excel.load_workbook(wb_path)
        tables = scan_workbook(workbook)
        df = extract_table(workbook, tables[0], schema)

        # Verify columns are normalized to lowercase canonical names
        assert "region" in df.columns
        assert "process" in df.columns

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_handles_empty_cells(schema):
    """Test that empty cells are properly handled as None."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Empty"

    sheet["A1"] = "~FI_T: EmptyTest"
    sheet["A2"] = "Region"
    sheet["B2"] = "Process"
    sheet["C2"] = "Year"

    # Row with empty cell
    sheet["A3"] = "AUS"
    sheet["B3"] = None  # Empty
    sheet["C3"] = 2020

    # Row with whitespace-only value
    sheet["A4"] = "NZ"
    sheet["B4"] = "   "
    sheet["C4"] = 2025

    wb_path = save_temp_workbook(wb)

    try:
        workbook = excel.load_workbook(wb_path)
        tables = scan_workbook(workbook)
        df = extract_table(workbook, tables[0], schema)

        # Verify None handling
        assert df.iloc[0]["process"] is None or pd.isna(df.iloc[0]["process"])
        assert df.iloc[1]["process"] is None or pd.isna(df.iloc[1]["process"])

        # Verify non-empty values are preserved
        assert df.iloc[0]["region"] == "AUS"
        assert df.iloc[0]["year"] == "2020"

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_preserves_row_order(schema):
    """Test that row order matches Excel (sorting happens later)."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Order"

    sheet["A1"] = "~FI_T: OrderTest"
    sheet["A2"] = "Region"
    sheet["B2"] = "Year"

    # Deliberately unordered data
    sheet["A3"] = "ZZZ"
    sheet["B3"] = 2030

    sheet["A4"] = "AAA"
    sheet["B4"] = 2010

    sheet["A5"] = "MMM"
    sheet["B5"] = 2020

    wb_path = save_temp_workbook(wb)

    try:
        workbook = excel.load_workbook(wb_path)
        tables = scan_workbook(workbook)
        df = extract_table(workbook, tables[0], schema)

        # Verify order matches Excel row order (not sorted)
        assert df.iloc[0]["region"] == "ZZZ"
        assert df.iloc[1]["region"] == "AAA"
        assert df.iloc[2]["region"] == "MMM"

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_multiple_tables(schema, multiple_tables_workbook):
    """Test extracting multiple tables from same workbook."""
    wb_path = save_temp_workbook(multiple_tables_workbook)

    try:
        # Scan to get all tables
        workbook = excel.load_workbook(wb_path)
        tables = scan_workbook(workbook)
        assert len(tables) == 3

        # Extract each table
        dfs = []
        for table_meta in tables:
            df = extract_table(workbook, table_meta, schema)
            dfs.append(df)

        # Verify all extracted correctly
        assert len(dfs) == 3

        # Table 1: 2 rows, 2 columns
        assert dfs[0].shape == (2, 2)

        # Table 2: 2 rows, 2 columns
        assert dfs[1].shape == (2, 2)

        # Table 3: 1 row, 2 columns
        assert dfs[2].shape == (1, 2)

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_sheet_not_found(schema):
    """Test error handling when sheet doesn't exist."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Exists"
    sheet["A1"] = "~FI_T: Test"
    sheet["A2"] = "Region"
    sheet["A3"] = "AUS"

    wb_path = save_temp_workbook(wb)

    try:
        # Create fake table_meta with non-existent sheet
        fake_meta = {
            "tag": "~FI_T: Test",
            "tag_type": "fi_t",
            "sheet_name": "DoesNotExist",
            "tag_row": 1,
            "tag_col": 1,
        }

        # Should raise ValueError
        workbook = excel.load_workbook(wb_path)
        with pytest.raises(ValueError, match="Sheet 'DoesNotExist' not found"):
            extract_table(workbook, fake_meta, schema)

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_whitespace_stripping(schema):
    """Test that whitespace is stripped from values."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Whitespace"

    sheet["A1"] = "~FI_T: WhitespaceTest"
    sheet["A2"] = "Region"
    sheet["B2"] = "Process"

    # Values with leading/trailing whitespace
    sheet["A3"] = "  AUS  "
    sheet["B3"] = "  COAL_PWR  "

    sheet["A4"] = "\tNZ\t"
    sheet["B4"] = "\nWIND_PWR\n"

    wb_path = save_temp_workbook(wb)

    try:
        workbook = excel.load_workbook(wb_path)
        tables = scan_workbook(workbook)
        df = extract_table(workbook, tables[0], schema)

        # Verify whitespace is stripped
        assert df.iloc[0]["region"] == "AUS"
        assert df.iloc[0]["process"] == "COAL_PWR"
        assert df.iloc[1]["region"] == "NZ"
        assert df.iloc[1]["process"] == "WIND_PWR"

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_numeric_precision_preserved(schema):
    """Test that numeric values are preserved as strings."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Numeric"

    sheet["A1"] = "~FI_T: NumericTest"
    sheet["A2"] = "Year"

    # Various numeric formats
    sheet["A3"] = 2020
    sheet["A4"] = 2025
    sheet["A5"] = 2030
    sheet["A6"] = 2035

    wb_path = save_temp_workbook(wb)

    try:
        workbook = excel.load_workbook(wb_path)
        tables = scan_workbook(workbook)
        df = extract_table(workbook, tables[0], schema)

        # All values should be strings
        for val in df["year"]:
            assert isinstance(val, str)

        # Check specific values (converted to string)
        assert df["year"].values[0] == "2020"
        assert df["year"].values[1] == "2025"

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_empty_table_no_data_rows(schema):
    """Test extracting table with headers but no data rows."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Empty"

    sheet["A1"] = "~FI_T: EmptyTable"
    sheet["A2"] = "Region"
    sheet["B2"] = "Process"
    # No data rows

    wb_path = save_temp_workbook(wb)

    try:
        workbook = excel.load_workbook(wb_path)
        tables = scan_workbook(workbook)
        assert len(tables) == 1

        df = extract_table(workbook, tables[0], schema)

        # Should have columns but no rows
        assert df.shape == (0, 2)
        assert list(df.columns) == ["region", "process"]

    finally:
        Path(wb_path).unlink(missing_ok=True)


def test_extract_with_fixture(schema):
    """Test extraction with real fixture workbook."""
    fixture_path = Path(__file__).parent / "fixtures" / "sample_deck" / "VT_BaseYear.xlsx"

    if not fixture_path.exists():
        pytest.skip(f"Fixture not found: {fixture_path}")

    # Load workbook and scan
    workbook = excel.load_workbook(str(fixture_path))
    tables = scan_workbook(workbook)

    if len(tables) == 0:
        pytest.skip("No tables found in fixture")

    # Find a table with headers (skip empty tables)
    df = None
    for table in tables:
        test_df = extract_table(workbook, table, schema)
        if len(test_df.columns) > 0:
            df = test_df
            break

    if df is None:
        pytest.skip("No non-empty tables found in fixture")

    # Basic validation
    assert df is not None
    assert len(df.columns) > 0

    # All values should be strings or None
    for col in df.columns:
        for val in df[col]:
            assert val is None or isinstance(val, str) or pd.isna(val)
