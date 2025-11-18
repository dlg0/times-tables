"""Tests for Excel extraction utilities."""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from austimes_tables.excel import (
    find_tags,
    get_sheet_names,
    hash_workbook,
    load_workbook,
    read_table_range,
)


@pytest.fixture
def sample_workbook():
    """Create a synthetic workbook with VEDA tags and tables."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "TestSheet"

    # Add a VEDA tag at B5
    sheet["B5"] = "~FI_T: BaseParams"

    # Add headers at row 6 (tag_row + 1)
    sheet["B6"] = "Region"
    sheet["C6"] = "Process"
    sheet["D6"] = "Value"

    # Add data rows
    sheet["B7"] = "AUS"
    sheet["C7"] = "COAL_PWR"
    sheet["D7"] = 100.5

    sheet["B8"] = "AUS"
    sheet["C8"] = "GAS_PWR"
    sheet["D8"] = 75.2

    # Add another tag at E10
    sheet["E10"] = "~TFM_INS: Commodities"
    sheet["E11"] = "Commodity"
    sheet["E12"] = "ELEC"

    # Add second sheet
    wb.create_sheet("EmptySheet")

    return wb


@pytest.fixture
def temp_workbook(sample_workbook):
    """Save sample workbook to temporary file."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    sample_workbook.save(tmp_path)
    yield tmp_path

    # Cleanup
    Path(tmp_path).unlink(missing_ok=True)


def test_load_workbook(temp_workbook):
    """Test loading workbook with correct settings."""
    wb = load_workbook(temp_workbook)
    assert wb is not None
    assert isinstance(wb, openpyxl.Workbook)
    assert len(wb.sheetnames) == 2


def test_get_sheet_names(sample_workbook):
    """Test retrieving sheet names."""
    names = get_sheet_names(sample_workbook)
    assert names == ["TestSheet", "EmptySheet"]


def test_find_tags(sample_workbook):
    """Test finding VEDA tags in sheet."""
    sheet = sample_workbook["TestSheet"]
    tags = find_tags(sheet)

    assert len(tags) == 2

    # First tag
    assert tags[0]["row"] == 5
    assert tags[0]["col"] == 2
    assert tags[0]["value"] == "~FI_T: BaseParams"
    assert tags[0]["cell_ref"] == "B5"

    # Second tag
    assert tags[1]["row"] == 10
    assert tags[1]["col"] == 5
    assert tags[1]["value"] == "~TFM_INS: Commodities"
    assert tags[1]["cell_ref"] == "E10"


def test_find_tags_empty_sheet(sample_workbook):
    """Test finding tags in empty sheet."""
    sheet = sample_workbook["EmptySheet"]
    tags = find_tags(sheet)
    assert tags == []


def test_read_table_range(sample_workbook):
    """Test reading table headers and data."""
    sheet = sample_workbook["TestSheet"]

    # Read first table (tag at row 5, col 2)
    headers, data_rows = read_table_range(sheet, start_row=5, start_col=2)

    # Check headers
    assert headers == ["Region", "Process", "Value"]

    # Check data rows
    assert len(data_rows) == 2
    assert data_rows[0] == ["AUS", "COAL_PWR", 100.5]
    assert data_rows[1] == ["AUS", "GAS_PWR", 75.2]


def test_read_table_range_single_column(sample_workbook):
    """Test reading single-column table."""
    sheet = sample_workbook["TestSheet"]

    # Read second table (tag at row 10, col 5)
    headers, data_rows = read_table_range(sheet, start_row=10, start_col=5)

    assert headers == ["Commodity"]
    assert len(data_rows) == 1
    assert data_rows[0] == ["ELEC"]


def test_read_table_range_empty(sample_workbook):
    """Test reading from position with no headers."""
    sheet = sample_workbook["EmptySheet"]

    headers, data_rows = read_table_range(sheet, start_row=1, start_col=1)

    assert headers == []
    assert data_rows == []


def test_hash_workbook(temp_workbook):
    """Test workbook hashing produces stable hash."""
    hash1 = hash_workbook(temp_workbook)
    hash2 = hash_workbook(temp_workbook)

    # Same file should produce same hash
    assert hash1 == hash2
    assert len(hash1) == 64  # SHA256 hex string length
    assert all(c in "0123456789abcdef" for c in hash1)


def test_hash_workbook_different_files(temp_workbook):
    """Test different files produce different hashes."""
    # Create another workbook
    wb2 = openpyxl.Workbook()
    wb2.active["A1"] = "Different content"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path2 = tmp.name

    wb2.save(tmp_path2)

    try:
        hash1 = hash_workbook(temp_workbook)
        hash2 = hash_workbook(tmp_path2)

        assert hash1 != hash2
    finally:
        Path(tmp_path2).unlink(missing_ok=True)
